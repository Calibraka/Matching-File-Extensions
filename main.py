import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

def find_matching_files(directory, extensions):
    folders_with_files = []

    for folder in os.listdir(directory):
        folder_path = os.path.join(directory, folder)
        if os.path.isdir(folder_path):
            matching_extensions = {ext: False for ext in extensions}
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_ext = os.path.splitext(file)[1].lower()
                    if file_ext in matching_extensions:
                        matching_extensions[file_ext] = True
                # Do not break here to check all subfolders
            folders_with_files.append((folder, matching_extensions))

    return folders_with_files

def write_to_excel(data, output_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Top Level Folders and Files'

    extensions = sorted(set(ext for (_, exts) in data for ext in exts.keys()))

    sheet['A1'] = 'Folder Name'
    for col_num, ext in enumerate(extensions, start=2):
        sheet.cell(row=1, column=col_num, value=ext)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    light_green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for row_num, (folder_name, matching_extensions) in enumerate(data, start=2):
        sheet.cell(row=row_num, column=1, value=folder_name).border = thin_border
        for col_num, ext in enumerate(extensions, start=2):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if matching_extensions[ext]:
                cell.value = '✓'
                cell.fill = light_green_fill
            else:
                cell.value = 'X'
                cell.fill = light_red_fill

    workbook.save(output_file)
    print(f"Data saved to '{output_file}'.")

if __name__ == "__main__":
    # Directory to scan
    scan_directory = 'Enter directory here'

    # Output file
    output_file = 'TEMP.xlsx'

    # Specific extensions to find
    extensions_to_find = ['.cad', '.stl', '.igs', '.dxf', '.dwg', '.top', '.atc']

    # Find top-level folders with matching files
    data = find_matching_files(scan_directory, extensions_to_find)

    # Write data to Excel
    write_to_excel(data, output_file)
